from bs4 import BeautifulSoup as bs4
from openpyxl import Workbook, load_workbook
import requests
import os
import lxml
import cchardet

BASE_URL = 'https://cataz.net/'
GENRE_URL = 'genre/horror'
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:100.0) Gecko/20100101 Firefox/100.0'
}
XLSX_FILE = 'movies.xlsx'


def get_html():
    r = requests.get(BASE_URL + GENRE_URL, headers=headers)
    return r


def get_total_page_num(html):
    soup = bs4(html.text, 'lxml')
    return int(soup.find('a', title='Last')['href'].split('=')[-1])


def get_all_movies_urls(pages_num):
    urls = []
    print('Getting urls from pages...')
    for i in range(1, pages_num + 1):
        if i == 1:
            r = requests.get(BASE_URL + GENRE_URL, headers=headers)
        else:
            params = {
                'page': i,
            }
            r = requests.get(BASE_URL + GENRE_URL, headers=headers, params=params)
        soup = bs4(r.text, 'html.parser')
        movies_urls = soup.find_all('a', {'class': 'film-poster-ahref flw-item-tip'})
        for data in movies_urls:
            urls.append(data['href'])
        get_downloading_bar(i, pages_num)
    print('\n')
    return urls


def get_downloading_bar(cur_num, total_num):
    load = int(cur_num / total_num * 100)
    unload = 100 - load
    answer = '\r[' + '#' * load + '-' * unload + f'] {cur_num}/{total_num}'
    print(answer, end='')


def get_movie_data(route):
    r = requests.get(BASE_URL + route, headers=headers)
    soup = bs4(r.text, 'lxml')
    name = soup.find('h2', class_='heading-name').a.contents[0]
    imdb = soup.find('span', class_='item mr-2').button.contents[0].split()[-1]
    year = soup.find('div', class_='col-xl-5 col-lg-6 col-md-8 col-sm-12')
    year = year.find_all('div', class_='row-line')[0].contents[2][1:5]
    genre = soup.find('div', class_='col-xl-5 col-lg-6 col-md-8 col-sm-12')
    genre = genre.find_all('div', class_='row-line')[1].find_all('a')
    genre = ' '.join([x['title'] for x in genre])
    duration = soup.find('div', class_='col-xl-6 col-lg-6 col-md-4 col-sm-12')
    duration = duration.find_all('div', class_='row-line')[0].contents[2].split()
    duration = duration[0] + ' ' + duration[1]
    country = soup.find('div', class_='col-xl-6 col-lg-6 col-md-4 col-sm-12')
    country = country.find_all('div', class_='row-line')[1].find_all('a')
    country = ' '.join([x['title'] for x in country])
    r.close()
    return (name, imdb, year, genre, duration, country, BASE_URL + route)


def write_in_excel(data):
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    for item in data:
        ws.append(item)
    wb.save(XLSX_FILE)
    wb.close()


def make_workbook():
    if os.path.exists(XLSX_FILE):
        os.remove(XLSX_FILE)
    wb = Workbook()
    ws = wb.active
    ws.append(('name', 'IMDB', 'year', 'genre', 'duration', 'country', 'URL'))
    ws.title = GENRE_URL.split('/')[-1]
    wb.save(XLSX_FILE)
    wb.close()


if __name__ == '__main__':
    make_workbook()
    pages = get_total_page_num(get_html())
    urls = get_all_movies_urls(pages)
    total_movies = len(urls)
    data = []
    errors = []
    counter = 1
    print('Parsing movies...')
    for url in urls:
        try:
            data.append(get_movie_data(url))
        except Exception as e:
            errors.append(f'#{counter}  {url} - {e.args[0]}')
        get_downloading_bar(counter, total_movies)
        if len(data) == 100:
            write_in_excel(data)
            data = []
        counter += 1
    print('\n')
    write_in_excel(data)
    if errors:
        print(*errors, sep='\n')
        print()
    print('Writing into xlsx complete!')
